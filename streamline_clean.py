import csv
from typing import NamedTuple
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
import jellyfish

class encampment(NamedTuple):
    tents: int
    structures: int
    vehicles: int

    year: int
    month: int
    date_time: datetime
    latitude: float
    longitude: float
    neighborhood: str
   

class encampment_report(NamedTuple):
    caseid: int
    year: int
    month: int
    date_time: datetime
    latitude: float
    longitude: float
    neighborhood: str
    status_notes: str
    report_ids: str

def rate(score):
    if score >= 0.95:
        return "high"
    if score < 0.95 and score >= 0.80:
        return "medium"
    return "low"   


DATA_DIR = Path(__file__).parent / "data"

## Clean 311 data
def clean_311():

    file_input = DATA_DIR / "311_cases.csv"

    with open(file_input, newline="") as csvfile:
        """
        Given a CSV containing 311, return a list of Encampment report objects.
        """
        output_report = []
        reader = csv.DictReader(csvfile)
        for row in reader:
            ### Clean the date ####
            datetime_str = row.get("Opened").replace(" PM", "")
            datetime_str = datetime_str.replace(" AM", "")
            datetime_object = datetime.strptime(datetime_str, "%m/%d/%Y %H:%M:%S")
            date_year = datetime_object.year
            date_month = datetime_object.month

            tuple_out = encampment_report(
                row.get("CaseID"),
                date_year,
                date_month,
                datetime_object,
                float(row.get("Latitude")),
                float(row.get("Longitude")),
                row.get("Neighborhood"),
                row.get("Status Notes").lower(),
                []
            )
            output_report.append(tuple_out)

    return output_report

### Clean encampment data ###
def clean_encampment():
    file_input = DATA_DIR / "Historical Tent Counts.xlsx"
    wb = load_workbook(file_input)
    sheet_obj = wb.active

    for i in range(1, sheet_obj.max_column + 1):
        print(sheet_obj.cell(row=1, column=i).value)

    for i in range(1, sheet_obj.max_column + 1):
        print(sheet_obj.cell(row=2, column=i).value)

    assert sheet_obj.cell(row=2, column=3).value == "Tents"
    assert sheet_obj.cell(row=2, column=4).value == "Structures"
    assert sheet_obj.cell(row=2, column=5).value == "Passenger Vehicles"
    assert sheet_obj.cell(row=2, column=6).value == 'Other Vehicles'
    assert sheet_obj.cell(row=2, column=8).value == 'Neighborhood'
    assert sheet_obj.cell(row=2, column=10).value == 'Latitude'
    assert sheet_obj.cell(row=2, column=11).value == 'Longitude'

    output_encampment = []
    for i in range(3, sheet_obj.max_row+1):

        sheet_obj.cell(row=3, column=1).value 
        date_obj = sheet_obj.cell(row=i, column=1).value
        date_string = date_obj.strftime("%m/%d/%Y")
        tents = sheet_obj.cell(row=i, column=3).value
        structure = sheet_obj.cell(row=i, column=4).value
        vehicles = sheet_obj.cell(row=i, column=5).value + sheet_obj.cell(row=i, column=6).value
        neighborhood = sheet_obj.cell(row=i, column=8).value

        lat = float(sheet_obj.cell(row=i, column=10).value)
        lon = float(sheet_obj.cell(row=i, column=11).value)
        obj = encampment(tents,
        structure,
        vehicles,
        date_obj.year,
        date_obj.month,
        date_string,
        lat,
        lon,
        neighborhood)
        output_encampment.append(obj)
    return output_encampment


#### Merge the two files to filter out 311 reports associatd with marked/observed encampments ####

def attached_311_reports(output_encampment, output_report):

    associated_encamp = []

    for encampment in output_encampment:
        for report in output_report: 
            if report.month == encampment.month and report.year == encampment.year:
                if rate(jellyfish.jaro_winkler_similarity(encampment.neighborhood.lower(), report.neighborhood.lower())) == "high":
                    associated_encamp.append((encampment, report))


 
